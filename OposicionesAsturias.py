import requests
from datetime import datetime, timedelta
import re, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import xml.etree.ElementTree as ET

# -------------------------------
# Función auxiliar: leer XML y devolver nota con "turno libre" o "promoción interna"
# -------------------------------
def obtener_nota_turno(url_xml):
    try:
        resp = requests.get(url_xml)
        if resp.status_code == 200:
            root = ET.fromstring(resp.content)
            notas = root.findall(".//notas/nota")
            for nota in notas:
                texto = "".join(nota.itertext()).strip()
                if "turno libre" in texto.lower() or "promoción interna" in texto.lower():
                    return texto  # devolvemos la nota completa
        return ""
    except Exception as e:
        print(f"⚠️ Error leyendo XML {url_xml}: {e}")
        return ""

# -------------------------------
# Función 1: Obtener JSON del BOE
# -------------------------------
def obtener_sumario(fecha_str):
    print(f"📥 Descargando sumario del BOE para la fecha {fecha_str}...")
    url = f"https://www.boe.es/datosabiertos/api/boe/sumario/{fecha_str}"
    headers = {"Accept": "application/json"}
    resp = requests.get(url, headers=headers)
    if resp.status_code == 200:
        print("✅ Sumario descargado correctamente.")
        return resp.json()
    else:
        print(f"⚠️ No se encontró BOE para {fecha_str} (código {resp.status_code}).")
    return None

# -----------------------------------
# Función 2: Extraer Ayuntamiento
# -----------------------------------
def extraer_ayuntamiento(titulo):
    patron = r"(Ayuntamiento[^()]+)"
    coincidencia = re.search(patron, titulo)
    return coincidencia.group(1).strip() if coincidencia else titulo

# -----------------------------------
# Función 3: Filtrar por sección y palabra clave + leer XML
# -----------------------------------
def filtrar_oposiciones_asturias(data, seccion_objetivo="B. Oposiciones y concursos", palabra_clave="Asturias"):
    resultados = []
    diarios = data["data"]["sumario"]["diario"]
    fecha_json = data["data"]["sumario"]["metadatos"]["fecha_publicacion"]
    fecha_formateada = datetime.strptime(fecha_json, "%Y%m%d").strftime("%d/%m/%Y")

    print(f"🔎 Analizando sumario del {fecha_formateada}...")

    for diario in diarios:
        for seccion in diario.get("seccion", []):
            if seccion_objetivo.lower() in seccion.get("nombre", "").lower():
                departamentos = seccion.get("departamento", [])
                if isinstance(departamentos, dict):
                    departamentos = [departamentos]

                for departamento in departamentos:
                    epigrafes = departamento.get("epigrafe", [])
                    if isinstance(epigrafes, dict):
                        epigrafes = [epigrafes]

                    for epigrafe in epigrafes:
                        items = epigrafe.get("item", [])
                        if isinstance(items, dict):
                            items = [items]

                        for item in items:
                            titulo = item.get("titulo", "")
                            enlace = item.get("url_html", "")
                            url_xml = item.get("url_xml", "")
                            if palabra_clave.lower() in titulo.lower():
                                ayuntamiento = extraer_ayuntamiento(titulo)

                                # --- Nuevo: leer XML y obtener nota relevante ---
                                nota_turno = ""
                                if url_xml:
                                    nota_turno = obtener_nota_turno(url_xml)

                                resultados.append([fecha_formateada, ayuntamiento, titulo, nota_turno, enlace])
    return resultados

# -----------------------------------
# Función 4: Buscar últimos N días
# -----------------------------------
def buscar_oposiciones_asturias(dias=15):
    print(f"📅 Buscando oposiciones y concursos de Asturias en los últimos {dias} días...\n")
    hoy = datetime.today()
    inicio = hoy - timedelta(days=dias)
    resultados_totales = []

    for i in range(dias+1):
        fecha = inicio + timedelta(days=i)
        fecha_str = fecha.strftime("%Y%m%d")
        data = obtener_sumario(fecha_str)
        if data:
            resultados = filtrar_oposiciones_asturias(data)
            resultados_totales.extend(resultados)

    return resultados_totales

# -----------------------------------
# Función 5: Mostrar tabla en consola
# -----------------------------------
def mostrar_tabla(resultados, dias):
    print("\n📊 RESULTADOS: Oposiciones y concursos en Asturias")
    print(f"(últimos {dias} días)\n")
    if resultados:
        print(f"{'Fecha':<12} | {'Ayuntamiento':<30} | {'Nota turno':<50} | {'Enlace'}")
        print("-"*160)
        for fila in resultados:
            nota_preview = (fila[3][:47] + "...") if len(fila[3]) > 50 else fila[3]
            print(f"{fila[0]:<12} | {fila[1]:<30} | {nota_preview:<50} | {fila[4]}")
        print(f"\n✅ Se encontraron {len(resultados)} convocatorias en el rango indicado.")
    else:
        print("❌ No se encontraron oposiciones y concursos de Asturias en el rango indicado.")

# -----------------------------------
# Función 6: Exportar a Excel
# -----------------------------------
def exportar_excel(resultados, nombre_archivo="oposiciones_asturias.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Convocatorias Asturias"

    encabezados = ["Fecha", "Ayuntamiento", "Título completo", "Nota turno", "Enlace"]
    ws.append(encabezados)

    for col_num, encabezado in enumerate(encabezados, 1):
        celda = ws.cell(row=1, column=col_num)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        celda.alignment = Alignment(horizontal="center", vertical="center")

    for fila in resultados:
        ws.append([fila[0], fila[1], fila[2], fila[3], ""])
        row_idx = ws.max_row
        enlace_celda = ws.cell(row=row_idx, column=5)
        enlace_celda.value = "Abrir BOE"
        enlace_celda.hyperlink = fila[4]
        enlace_celda.font = Font(color="0000FF", underline="single")

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    escritorio = os.path.join(os.path.expanduser("~"), "Desktop")
    ruta_final = os.path.join(escritorio, nombre_archivo)
    wb.save(ruta_final)
    print(f"📂 Datos exportados a Excel en la ruta {ruta_final}")

# -------------------------------
# Ejecución principal
# -------------------------------
if __name__ == "__main__":
    dias = 15
    resultados = buscar_oposiciones_asturias(dias=dias)
    mostrar_tabla(resultados, dias)

    if resultados:
        opcion = input("\n¿Desea exportar esta información a Excel? (s/n): ").strip().lower()
        if opcion == "s":
            exportar_excel(resultados)
        else:
            print("ℹ️ Exportación cancelada.")
